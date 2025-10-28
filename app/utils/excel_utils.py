import openpyxl
from typing import Dict, List


def extract_data_from_excel(file_path: str) -> Dict[str, List[List]]:
    """
    Extract data from an Excel file.

    Args:
        file_path: Path to the Excel file

    Returns:
        dict: Dictionary with sheet names as keys and data as values
    """
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    data = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_data = []

        for row in sheet.iter_rows(values_only=True):
            # Filter out completely empty rows
            if any(cell is not None for cell in row):
                sheet_data.append(list(row))

        data[sheet_name] = sheet_data

    workbook.close()
    return data


def excel_to_text(file_path: str) -> str:
    """
    Convert Excel file content to text format.

    Args:
        file_path: Path to the Excel file

    Returns:
        str: Text representation of the Excel content
    """
    data = extract_data_from_excel(file_path)
    text_parts = []

    for sheet_name, sheet_data in data.items():
        text_parts.append(f"Sheet: {sheet_name}")
        for row in sheet_data:
            row_text = "\t".join(str(cell) if cell is not None else "" for cell in row)
            text_parts.append(row_text)
        text_parts.append("")  # Empty line between sheets

    return "\n".join(text_parts)
